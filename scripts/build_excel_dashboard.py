"""Phase 3: build the Excel dashboard from validated dashboard source tables."""

from __future__ import annotations

import csv
import math
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_DIR = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = PROJECT_DIR / "excel_dashboard/patient_discharge_experience_dashboard.xlsx"
BACKUP_PATH = PROJECT_DIR / "excel_dashboard/patient_discharge_experience_dashboard.phase3_backup.xlsx"
BUILD_NOTES = PROJECT_DIR / "docs/excel_dashboard_build_notes.md"

SOURCE_FILES = {
    "category_cards": PROJECT_DIR / "data/processed/dashboard_category_cards.csv",
    "category_comparison": PROJECT_DIR / "data/processed/dashboard_category_comparison.csv",
    "submeasure_breakdown": PROJECT_DIR / "data/processed/dashboard_submeasure_breakdown.csv",
    "priority_findings": PROJECT_DIR / "data/processed/dashboard_priority_findings.csv",
    "hospital_comparison": PROJECT_DIR / "data/processed/dashboard_hospital_comparison.csv",
}

COLORS = {
    "navy": "155E75",
    "teal": "0F766E",
    "blue": "2563EB",
    "gray": "F3F6F8",
    "pale_blue": "EAF3F8",
    "pale_teal": "E6F4F1",
    "orange": "F97316",
    "pale_orange": "FFF2E6",
    "white": "FFFFFF",
    "text": "1F2937",
    "muted": "64748B",
    "border": "D6E0E5",
}

INVALID_STRINGS = {
    "nan",
    "inf",
    "infinity",
    "-inf",
    "-infinity",
    "none",
    "#div/0!",
    "#value!",
    "#num!",
    "#n/a",
    "#ref!",
    "#name?",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as source:
        return list(csv.DictReader(source))


def num(value: object, divisor: float = 1.0) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        parsed = float(str(value).strip()) / divisor
    except ValueError:
        return None
    return parsed if math.isfinite(parsed) else None


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    if text.strip().lower() in INVALID_STRINGS:
        return ""
    return text


def percent_display(value: object, digits: int = 1) -> str:
    parsed = num(value)
    if parsed is None:
        return ""
    return f"{parsed:.{digits}f}%"


def set_cell(ws, row: int, col: int, value: object, style: str | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    if style == "header":
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    elif style == "section":
        cell.fill = PatternFill("solid", fgColor=COLORS["teal"])
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_box(ws, cell_range: str, fill: str = "white", font_color: str = "text") -> None:
    side = Side(style="thin", color=COLORS["border"])
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=COLORS[fill])
            cell.font = Font(color=COLORS[font_color])
            cell.border = Border(left=side, right=side, top=side, bottom=side)
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def recreate_sheet(wb, name: str, index: int | None = None):
    if name in wb.sheetnames:
        old = wb[name]
        idx = wb.sheetnames.index(name)
        wb.remove(old)
        return wb.create_sheet(name, idx)
    if index is None:
        return wb.create_sheet(name)
    return wb.create_sheet(name, index)


def configure_sheet(ws, widths: dict[str, float]) -> None:
    ws.sheet_view.showGridLines = False
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_table(ws, name: str, start_row: int, start_col: int, headers: list[str], rows: list[list[object]]) -> tuple[int, int]:
    for col_offset, header in enumerate(headers):
        set_cell(ws, start_row, start_col + col_offset, header, "header")
    for row_offset, row in enumerate(rows, start=1):
        for col_offset, value in enumerate(row):
            set_cell(ws, start_row + row_offset, start_col + col_offset, value)
    end_row = start_row + len(rows)
    end_col = start_col + len(headers) - 1
    ref = f"{ws.cell(start_row, start_col).coordinate}:{ws.cell(end_row, end_col).coordinate}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    apply_box(ws, ref)
    for cell in ws[start_row]:
        if start_col <= cell.column <= end_col:
            cell.fill = PatternFill("solid", fgColor=COLORS["teal"])
            cell.font = Font(color=COLORS["white"], bold=True)
    return end_row, end_col


def add_title(ws, title: str, subtitle: str | None = None, note: str | None = None, end_col: str = "N") -> None:
    ws.merge_cells(f"B2:{end_col}3")
    ws["B2"] = title
    ws["B2"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["B2"].font = Font(color=COLORS["white"], bold=True, size=20)
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if subtitle:
        ws.merge_cells(f"B4:{end_col}4")
        ws["B4"] = subtitle
        ws["B4"].fill = PatternFill("solid", fgColor=COLORS["pale_blue"])
        ws["B4"].font = Font(color=COLORS["text"], bold=True)
        ws["B4"].alignment = Alignment(horizontal="center")
    if note:
        ws.merge_cells(f"B5:{end_col}5")
        ws["B5"] = note
        ws["B5"].fill = PatternFill("solid", fgColor=COLORS["pale_blue"])
        ws["B5"].font = Font(color=COLORS["muted"])
        ws["B5"].alignment = Alignment(horizontal="center", wrap_text=True)


def build_dashboard_source(wb, data) -> None:
    ws = recreate_sheet(wb, "Dashboard_Source")
    configure_sheet(ws, {"A": 24, "B": 20, "C": 20, "D": 28, "E": 14, "G": 26, "H": 18, "I": 18, "J": 16, "K": 16, "L": 18, "M": 16, "N": 46})
    ws.merge_cells("A1:N1")
    set_cell(ws, 1, 1, "Dashboard Source Tables", "header")
    ws["A1"].font = Font(color=COLORS["white"], bold=True, size=18)
    ws.merge_cells("A2:N2")
    ws["A2"] = "Validated Phase 2 CSV sources loaded for dashboard auditability."
    ws["A2"].fill = PatternFill("solid", fgColor=COLORS["pale_blue"])

    add_table(
        ws,
        "tblCategoryCards",
        5,
        1,
        ["metric_name", "weighted_favorable_percent", "display_value", "interpretation_label", "sort_order"],
        [[r["metric_name"], num(r["weighted_favorable_percent"], 100), r["display_value"], r["interpretation_label"], num(r["sort_order"])] for r in data["category_cards"]],
    )
    add_table(
        ws,
        "tblCategoryComparison",
        5,
        7,
        ["category", "weighted_favorable_percent", "simple_average", "favorable_row_count", "facility_count", "total_survey_weight", "rank", "interpretation_note"],
        [[r["category"], num(r["weighted_favorable_percent"], 100), num(r["simple_average"], 100), num(r["favorable_row_count"]), num(r["facility_count"]), num(r["total_survey_weight"]), num(r["rank"]), r["interpretation_note"]] for r in data["category_comparison"]],
    )
    add_table(
        ws,
        "tblSubmeasureBreakdown",
        15,
        1,
        ["category", "submeasure_label", "favorable_response_definition", "weighted_favorable_percent", "simple_average", "favorable_row_count", "total_survey_weight", "interpretation_note"],
        [[r["category"], r["submeasure_label"], r["favorable_response_definition"], num(r["weighted_favorable_percent"], 100), num(r["simple_average"], 100), num(r["favorable_row_count"]), num(r["total_survey_weight"]), r["interpretation_note"]] for r in data["submeasure_breakdown"]],
    )
    add_table(
        ws,
        "tblPriorityFindings",
        35,
        1,
        ["finding_title", "metric_value", "evidence", "stakeholder_interpretation", "recommended_followup"],
        [[r["finding_title"], num(r["metric_value"], 100), r["evidence"], r["stakeholder_interpretation"], r["recommended_followup"]] for r in data["priority_findings"]],
    )
    add_table(
        ws,
        "tblHospitalComparison",
        45,
        1,
        ["facility_id", "facility_name", "category", "weighted_favorable_percent", "response_count", "response_rate_context", "florida_category_average", "difference_from_florida_average"],
        [[r["facility_id"], r["facility_name"], r["category"], num(r["weighted_favorable_percent"], 100), num(r["response_count"]), num(r["response_rate_context"], 100) if r["response_rate_context"] else "Not available", num(r["florida_category_average"], 100), num(r["difference_from_florida_average"])] for r in data["hospital_comparison"]],
    )
    for col in ["B", "D", "F", "G", "H", "I"]:
        for cell in ws[col]:
            if isinstance(cell.value, float) and 0 <= cell.value <= 1:
                cell.number_format = "0.0%"
    ws.freeze_panes = "A5"


def build_dashboard(wb, data) -> None:
    ws = recreate_sheet(wb, "Dashboard")
    configure_sheet(ws, {col: width for col, width in zip("ABCDEFGHIJKLMN", [4, 19, 19, 19, 19, 19, 19, 19, 19, 22, 22, 22, 22, 22])})
    for row in range(1, 45):
        ws.row_dimensions[row].height = 24
    add_title(
        ws,
        "Patient Discharge Experience & Follow-Up Support Survey Insights",
        "Florida CMS HCAHPS hospital survey data, 07/01/2024-06/30/2025",
        "Public hospital-level CMS HCAHPS results; Qualtrics follow-up survey design comes later.",
    )

    card_ranges = ["B7:C10", "D7:E10", "F7:G10", "H7:I10", "J7:K10", "L7:M10"]
    for row, cell_range in zip(data["category_cards"], card_ranges):
        ws.merge_cells(cell_range)
        top_left = cell_range.split(":")[0]
        ws[top_left] = f"{row['metric_name']}\n{percent_display(row['weighted_favorable_percent']) }\n{row['interpretation_label']}"
        priority = row["metric_name"] == "Medicine Communication"
        apply_box(ws, cell_range, "pale_orange" if priority else "white", "orange" if priority else "text")
        ws[top_left].font = Font(color=COLORS["orange" if priority else "text"], bold=True, size=12)
        ws[top_left].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("B12:H12")
    set_cell(ws, 12, 2, "Category comparison (% favorable)", "section")
    sorted_categories = sorted(data["category_comparison"], key=lambda r: num(r["weighted_favorable_percent"]) or 0, reverse=True)
    add_table(
        ws,
        "tblDashboardCategoryChart",
        13,
        2,
        ["Category", "Weighted favorable %"],
        [[r["category"], num(r["weighted_favorable_percent"], 100)] for r in sorted_categories],
    )
    for cell in ws["C"]:
        if isinstance(cell.value, float):
            cell.number_format = "0.0%"
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Weighted favorable percent by category"
    chart.y_axis.title = "Category"
    chart.x_axis.title = "% favorable"
    chart.x_axis.numFmt = "0%"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    data_ref = Reference(ws, min_col=3, min_row=13, max_row=19)
    cats_ref = Reference(ws, min_col=2, min_row=14, max_row=19)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 8
    chart.width = 15
    ws.add_chart(chart, "E14")

    ws.merge_cells("J12:N12")
    set_cell(ws, 12, 10, "Priority improvement callout", "section")
    callout = "\n".join([f"• {r['finding_title']}: {r['evidence']}" for r in data["priority_findings"][:2]])
    ws.merge_cells("J13:N20")
    ws["J13"] = f"{callout}\n• Side-effect explanation should be the main Qualtrics follow-up focus later."
    apply_box(ws, "J13:N20", "pale_orange")

    ws.merge_cells("J22:N25")
    ws["J22"] = "Discharge information is comparatively strong. Nurse and doctor communication are moderately strong, but explanation/listening trail respect/courtesy. Medicine communication is the clearest improvement area, especially side-effect explanation."
    apply_box(ws, "J22:N25", "pale_teal")
    ws.merge_cells("J26:N29")
    ws["J26"] = "The follow-up Qualtrics survey should focus on medication purpose, side-effect clarity, confidence following medication instructions, knowing who to contact, and unresolved medication concerns after discharge."
    apply_box(ws, "J26:N29", "pale_blue")

    ws.merge_cells("B30:N30")
    set_cell(ws, 30, 2, "Submeasure breakdown", "section")
    sub_rows = [
        [r["category"], r["submeasure_label"], r["favorable_response_definition"], num(r["weighted_favorable_percent"], 100), num(r["simple_average"], 100), num(r["favorable_row_count"]), num(r["total_survey_weight"])]
        for r in data["submeasure_breakdown"]
    ]
    add_table(ws, "tblDashboardSubmeasureBreakdown", 31, 2, ["Category", "Submeasure", "Definition", "Weighted %", "Simple avg", "Rows", "Survey weight"], sub_rows)
    for row in ws.iter_rows(min_row=32, max_row=47, min_col=5, max_col=6):
        for cell in row:
            cell.number_format = "0.0%"


def build_category_detail(wb, data) -> None:
    ws = recreate_sheet(wb, "Category_Detail")
    configure_sheet(ws, {"A": 24, "B": 30, "C": 22, "D": 18, "E": 16, "F": 12, "G": 16, "H": 48, "J": 18, "K": 18})
    add_title(ws, "Category and Submeasure Detail", "Submeasure-level favorable-response results", "Read each row as the survey-weighted favorable percent for a specific HCAHPS submeasure.", "H")
    rows = [[r["category"], r["submeasure_label"], r["favorable_response_definition"], num(r["weighted_favorable_percent"], 100), num(r["simple_average"], 100), num(r["favorable_row_count"]), num(r["total_survey_weight"]), r["interpretation_note"]] for r in data["submeasure_breakdown"]]
    add_table(ws, "tblCategoryDetailSubmeasures", 7, 1, ["Category", "Submeasure", "Favorable response", "Weighted favorable %", "Simple average", "Rows", "Survey weight", "Interpretation note"], rows)
    for row in ws.iter_rows(min_row=8, max_row=23, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = "0.0%"
    ws.conditional_formatting.add("D8:D23", ColorScaleRule(start_type="min", start_color="F97316", mid_type="percentile", mid_value=50, mid_color="FDE68A", end_type="max", end_color="0F766E"))
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Submeasure favorable percent"
    chart.x_axis.numFmt = "0%"
    data_ref = Reference(ws, min_col=4, min_row=7, max_row=23)
    cats_ref = Reference(ws, min_col=2, min_row=8, max_row=23)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 10
    chart.width = 16
    ws.add_chart(chart, "J7")
    ws.freeze_panes = "A8"


def build_hospital_detail(wb, data) -> None:
    ws = recreate_sheet(wb, "Hospital_Detail")
    configure_sheet(ws, {"A": 14, "B": 36, "C": 24, "D": 18, "E": 16, "F": 22, "G": 18, "H": 18, "J": 22, "K": 34})
    add_title(ws, "Hospital Detail", "Facility-level favorable metrics and Florida average comparison", "Interpret facility-level differences cautiously because survey counts vary and HCAHPS is hospital-level public reporting data.", "H")
    ws["J2"] = "Category selector"
    ws["J3"] = "Hospital selector"
    ws["K2"] = "Medicine Communication"
    ws["K3"] = data["hospital_comparison"][0]["facility_name"]
    categories = sorted({r["category"] for r in data["hospital_comparison"]})
    hospitals = sorted({r["facility_name"] for r in data["hospital_comparison"]})
    cat_validation = DataValidation(type="list", formula1=f'"{",".join(categories)}"', allow_blank=True)
    ws.add_data_validation(cat_validation)
    cat_validation.add(ws["K2"])
    hidden = recreate_sheet(wb, "Validation_Lists")
    hidden.sheet_state = "hidden"
    for idx, hospital in enumerate(hospitals, start=1):
        hidden.cell(row=idx, column=1, value=hospital)
    hosp_validation = DataValidation(type="list", formula1=f"'Validation_Lists'!$A$1:$A${len(hospitals)}", allow_blank=True)
    ws.add_data_validation(hosp_validation)
    hosp_validation.add(ws["K3"])
    rows = [[r["facility_id"], r["facility_name"], r["category"], num(r["weighted_favorable_percent"], 100), num(r["response_count"]), num(r["response_rate_context"], 100) if r["response_rate_context"] else "Not available", num(r["florida_category_average"], 100), num(r["difference_from_florida_average"])] for r in data["hospital_comparison"]]
    add_table(ws, "tblHospitalDetailComparison", 7, 1, ["Facility ID", "Facility name", "Category", "Weighted favorable %", "Response count", "Response rate context", "Florida avg", "Diff from FL avg"], rows)
    for row in ws.iter_rows(min_row=8, max_row=7 + len(rows), min_col=4, max_col=4):
        for cell in row:
            cell.number_format = "0.0%"
    for row in ws.iter_rows(min_row=8, max_row=7 + len(rows), min_col=7, max_col=7):
        for cell in row:
            cell.number_format = "0.0%"
    for row in ws.iter_rows(min_row=8, max_row=7 + len(rows), min_col=6, max_col=6):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"
    for cell in ws["H"]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00"
    ws.freeze_panes = "A8"


def build_readme(wb) -> None:
    ws = recreate_sheet(wb, "README")
    configure_sheet(ws, {"A": 30, "B": 110})
    rows = [
        ["Workbook", "Patient Discharge Experience & Follow-Up Support Survey Insights Dashboard"],
        ["Data source", "Public CMS HCAHPS hospital-level patient survey data."],
        ["Scope", "Florida hospitals only."],
        ["Reporting period", "07/01/2024-06/30/2025."],
        ["Favorable-response metric definition", "Top-box favorable responses: Yes for discharge information, Always for nurse/doctor/medicine communication, 9 or 10 for overall rating, and Definitely yes for recommendation."],
        ["Survey-volume weighting", "Weighted favorable percent = sum(favorable percent * completed-survey count) / sum(completed-survey count), excluding invalid or zero denominators."],
        ["HCAHPS note", "HCAHPS is public hospital-level reporting data, not individual patient survey microdata."],
        ["Qualtrics note", "Qualtrics is a future follow-up survey design informed by HCAHPS findings; it is not the source of HCAHPS data."],
    ]
    add_table(ws, "tblReadme", 1, 1, ["Topic", "Description"], rows)


def build_insights_brief(wb) -> None:
    ws = recreate_sheet(wb, "Insights_Brief")
    configure_sheet(ws, {"A": 34, "B": 110})
    rows = [
        ["Discharge Information is strongest", "Discharge Information is strongest at 84.27% favorable."],
        ["Nurse/Doctor communication are moderately strong", "Nurse Communication is 76.39% and Doctor Communication is 74.82% favorable."],
        ["Medicine Communication is weakest", "Medicine Communication is weakest at 58.01% favorable."],
        ["Side-effect explanation is lowest", "Side-effect explanation is the lowest specific measure at 43.78% favorable."],
        ["Recommendation and Overall Rating are moderate", "Recommendation is 70.01% and Overall Rating is 69.07%."],
        ["Recommended follow-up", "Design the Qualtrics survey around medication clarity, side-effect explanation, confidence following medication instructions, knowing who to contact, and unresolved medication concerns after discharge."],
    ]
    add_table(ws, "tblInsightsBrief", 1, 1, ["Finding", "Brief"], rows)


def invalid_cells(wb) -> list[str]:
    bad: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if isinstance(value, float) and not math.isfinite(value):
                    bad.append(f"{ws.title}!{cell.coordinate}")
                if isinstance(value, str) and value.strip().lower() in INVALID_STRINGS:
                    bad.append(f"{ws.title}!{cell.coordinate}")
    return bad


def write_build_notes(validation: dict[str, object], data) -> None:
    kpis = "\n".join(f"- {row['metric_name']}: {row['weighted_favorable_percent']}%" for row in data["category_cards"])
    notes = f"""# Excel Dashboard Build Notes

## Sheets Created/Updated
- Dashboard_Source
- Dashboard
- Category_Detail
- Hospital_Detail
- README
- Insights_Brief
- Validation_Lists (hidden selector support)

## Source Tables Used
{chr(10).join(f"- `{path.relative_to(PROJECT_DIR)}`" for path in SOURCE_FILES.values())}

## Charts/Tables Added
- Dashboard_Source: Excel tables for KPI cards, category comparison, submeasure breakdown, priority findings, and hospital comparison.
- Dashboard: KPI cards, category comparison bar chart, priority callout, submeasure table, interpretation box, and Qualtrics next-step box.
- Category_Detail: submeasure table, conditional formatting, and bar chart.
- Hospital_Detail: hospital/facility comparison table, Florida average comparison, difference from average, and category/facility selectors.

## Final KPI Values
{kpis}

## Limitations
- No fake monthly trends were created because the HCAHPS data is a reporting-period extract.
- response_rate_context is populated where the cleaned HCAHPS source includes a facility response rate; missing facility rates display as `Not available`.
- Facility-level differences use public hospital-level reporting data and should not be interpreted as patient-level microdata.
- Screenshot references were used only for broad clean healthcare dashboard styling inspiration.

## Validation Result
- openpyxl reopened workbook successfully: {validation["reopened"]}.
- Invalid token/value scan count: {validation["invalid_count"]}.
- Required sheets present: {validation["required_sheets_present"]}.
- Sheet dimensions: {validation["dimensions"]}.
"""
    BUILD_NOTES.write_text(notes, encoding="utf-8")


def main() -> None:
    data = {name: read_csv(path) for name, path in SOURCE_FILES.items()}
    shutil.copy2(WORKBOOK_PATH, BACKUP_PATH)
    wb = load_workbook(WORKBOOK_PATH)

    build_dashboard_source(wb, data)
    build_dashboard(wb, data)
    build_category_detail(wb, data)
    build_hospital_detail(wb, data)
    build_readme(wb)
    build_insights_brief(wb)

    bad_before_save = invalid_cells(wb)
    if bad_before_save:
        shutil.copy2(BACKUP_PATH, WORKBOOK_PATH)
        raise RuntimeError(f"Invalid values before save: {bad_before_save[:10]}")

    wb.save(WORKBOOK_PATH)

    try:
        reopened = load_workbook(WORKBOOK_PATH, data_only=False)
        bad_after_save = invalid_cells(reopened)
        required = [
            "Clean_Data",
            "README",
            "Data_Dictionary",
            "Pivot_Source",
            "Dashboard",
            "Insights_Brief",
            "Dashboard_Source",
            "Category_Detail",
            "Hospital_Detail",
        ]
        validation = {
            "reopened": True,
            "invalid_count": len(bad_after_save),
            "required_sheets_present": all(sheet in reopened.sheetnames for sheet in required),
            "dimensions": {sheet: reopened[sheet].calculate_dimension() for sheet in required if sheet in reopened.sheetnames},
        }
        if bad_after_save or not validation["required_sheets_present"]:
            shutil.copy2(BACKUP_PATH, WORKBOOK_PATH)
            raise RuntimeError(f"Workbook validation failed: {validation}")
        write_build_notes(validation, data)
    except Exception:
        shutil.copy2(BACKUP_PATH, WORKBOOK_PATH)
        raise

    print("Workbook saved and validation passed")
    print(f"Build notes: {BUILD_NOTES.relative_to(PROJECT_DIR)}")


if __name__ == "__main__":
    main()
